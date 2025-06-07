from typing import Union, Optional, Iterable
import openpyxl
import openpyxl.reader
import openpyxl.reader.excel
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
from base_processor import base_processor

"""
该类是 openpyxl 库的实现, 支持对 xlsx 文件进行读写。

注意事项：
1. openpyxl 库的行列索引是从 1 开始的，而不是从 0 开始的。
2. openpyxl 库的以工作表名称索引工作表，而不是序号。
"""

class openpyxl_backend(base_processor):
    def __init__(self, book_path: str):
        super().__init__(book_path)
        self.wb = openpyxl.load_workbook(book_path, read_only=False, rich_text=False)

    def _sheet_index_to_worksheet(self, sheet_index):
        """
        根据工作表的索引返回对应的工作表对象。
    
        :param sheet_index: 工作表的索引，从0开始计数。
        :type sheet_index: int 
    
        返回:
        worksheet: 对应索引的工作表对象。
        """
        # 通过索引获取工作表名称，并返回对应的工作表对象
        return self.wb[self.wb.sheetnames[sheet_index]]
    
    def _save(self, book_path):
        try:
            self.wb.save(book_path)
        except:
            return False
        return True

    def _get_sheets(self):
        return tuple(self.wb.sheetnames)

    def _create_sheet(self, sheet_name, index):
        create_ws = self.wb.create_sheet(sheet_name, index)
        if isinstance(create_ws, openpyxl.worksheet.worksheet.Worksheet):
            return True
        else:
            return False
        
    def _max_sheet(self):
        return len(self.wb.sheetnames) - 1
    
    def _max_row(self, sheet_index):
        return self._sheet_index_to_worksheet(sheet_index).max_row
    
    def _max_column(self, sheet_index):
        return self._sheet_index_to_worksheet(sheet_index).max_column
        
    def _get_row(self, sheet_index, row, start_col = None, end_col = None):
        worksheet = self._sheet_index_to_worksheet(sheet_index)
        res = []
        for row in worksheet.iter_rows(row, row, start_col, end_col):
            for cell in row:
                res.append(cell.value)
        return res
    def _get_col(self, sheet_index, col, start_row = None, end_row = None):
        worksheet = self._sheet_index_to_worksheet(sheet_index)
        res = []
        for col in worksheet.iter_cols(col, col, start_row, end_row):
            for cell in col:
                res.append(cell.value)

        return res
    
    def _write_row(self, sheet_index, row, data, start_col = None, end_col = None):
        worksheet = self._sheet_index_to_worksheet(sheet_index)

        for cell in worksheet.iter_rows(row, row, start_col, end_col):
            for i in range(len(data)):
                cell[i].value = data[i]
        return True
    
    def _write_col(self, sheet_index, col, data, start_row = None, end_row = None):
        worksheet = self._sheet_index_to_worksheet(sheet_index)

        for cell in worksheet.iter_rows(start_row, end_row, col, col):
            for i in range(len(data)):
                cell[i].value = data[i]
        return True

if __name__ == "__main__":
    excel = openpyxl_backend("学期成绩导入模板.xls")
    sheets = excel.get_sheets()
    print(sheets)
    print(excel.get_data(1))
    print(excel.get_row(1, 2))

"""TODO:
"""